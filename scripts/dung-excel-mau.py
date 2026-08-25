#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Dựng file Excel mẫu từ data/tthc.json, đúng lược đồ mà
scripts/trich-xuat-du-lieu.py mong đợi.

Vì sao cần: file Excel thật của đơn vị không nằm trong Git (xem
data/source/README.md), nên nếu không có công cụ này thì không ai kiểm thử được
pipeline trích xuất - phần mà đơn vị dùng thường xuyên nhất.

Cách dùng - chạy khứ hồi để kiểm chứng pipeline:

    python3 scripts/dung-excel-mau.py . /tmp/mau.xlsx
    python3 scripts/trich-xuat-du-lieu.py /tmp/mau.xlsx --ngay-xuat 06/08/2026
    git diff --stat data/

`git diff` phải sạch, trừ 7 bản ghi có lĩnh vực rỗng: bộ trích xuất bản mới
chuẩn hoá chúng thành "Chưa phân loại", còn data/tthc.json trong kho được sinh
bằng bản cũ chưa có bước chuẩn hoá đó. Website hiển thị giống nhau ở cả hai
trường hợp vì frontend cũng có sẵn bước dự phòng tương tự.

CẢNH BÁO: đây là công cụ kiểm thử. File nó sinh ra KHÔNG phải dữ liệu nguồn -
tuyệt đối không dùng để cập nhật data/, luôn dùng file Excel chính thức của
đơn vị.
"""
import json, sys
from pathlib import Path
import openpyxl

goc = Path(sys.argv[1]); ra = Path(sys.argv[2])
ds = json.loads((goc / "data" / "tthc.json").read_text(encoding="utf-8"))

wb = openpyxl.Workbook(); wb.remove(wb.active)

# --- TTHC2-CAPXA: dữ liệu bắt đầu từ dòng 4 ---
ws = wb.create_sheet("TTHC2-CAPXA")
ws.append(["Danh mục thủ tục hành chính cấp xã"])   # dòng 1: tiêu đề
ws.append([])                                        # dòng 2: trống
ws.append(["STT", "Mã TTHC", "Tên TTHC", "Link truy cập", "Số quyết định",
           "Ngày quyết định", "Cơ quan ban hành", "Lĩnh vực", "Cấp thực hiện",
           "Đối tượng thực hiện", "Loại TTHC", "Từ khoá", "Địa chỉ tiếp nhận",
           "Cơ quan thực hiện", "Cơ quan có thẩm quyền", "Cơ quan phối hợp",
           "Cơ quan được uỷ quyền", "Mô tả", "Trạng thái"])  # dòng 3: đầu cột
COT = ["stt", "ma_tthc", "ten_tthc", "link_truy_cap", "so_quyet_dinh",
       "ngay_quyet_dinh", "co_quan_ban_hanh", "linh_vuc", "cap_thuc_hien",
       "doi_tuong_thuc_hien", "loai_tthc", "tu_khoa", "dia_chi_tiep_nhan",
       "co_quan_thuc_hien", "co_quan_co_tham_quyen", "co_quan_phoi_hop",
       "co_quan_duoc_uy_quyen", "mo_ta", "trang_thai"]
for t in ds:
    ws.append([t.get(c, "") for c in COT])

def sheet_con(ten, cot_truoc_ma, khoa, truong):
    """Sheet phụ: mã TTHC ở cột B (chỉ số 1), dữ liệu bắt đầu dòng 2."""
    w = wb.create_sheet(ten)
    w.append(["STT", "Mã TTHC"] + [""] * cot_truoc_ma + truong)
    for t in ds:
        for muc in t.get(khoa, []) or []:
            w.append(["", t["ma_tthc"]] + [""] * cot_truoc_ma
                     + [muc.get(f, "") for f in truong])

sheet_con("Cách thức thực hiện", 1, "cach_thuc_thuc_hien",
          ["hinh_thuc_nop", "thoi_gian", "phi_le_phi", "mo_ta"])
sheet_con("Trình tự thực hiện", 1, "trinh_tu_thuc_hien",
          ["truong_hop_xu_ly", "mo_ta_chi_tiet"])
sheet_con("TPHS", 1, "tphs",
          ["ma_tphs", "ten_giay_to", "loai_giay_to", "co_quan_ban_hanh",
           "ban_chinh", "ban_sao", "bm_dien_tu"])
sheet_con("KQTH", 1, "kqth",
          ["ma_ket_qua", "ten_ket_qua", "co_quan_ban_hanh", "mo_ta"])
sheet_con("CCPL", 1, "ccpl",
          ["so_ky_hieu", "trich_yeu", "co_quan_ban_hanh", "ngay_ban_hanh",
           "ngay_hieu_luc", "dia_chi_truy_cap"])

# --- TTHC: yêu cầu điều kiện ở cột chỉ số 16, mã ở chỉ số 1, từ dòng 2 ---
w = wb.create_sheet("TTHC")
w.append(["STT", "Mã TTHC"] + [""] * 14 + ["Yêu cầu, điều kiện"])
# File thật của đơn vị có dòng cho mọi mã, kể cả khi ô yêu cầu-điều kiện rỗng.
# Bỏ dòng rỗng đi sẽ khiến khoá biến mất khỏi JSON thay vì mang giá trị "".
for t in ds:
    if "yeu_cau_dieu_kien" in t:
        w.append(["", t["ma_tthc"]] + [""] * 14 + [t["yeu_cau_dieu_kien"]])

wb.save(ra)
print(f"Đã ghi {ra}  ({ra.stat().st_size/1024:.0f} KB, {len(ds)} TTHC)")
