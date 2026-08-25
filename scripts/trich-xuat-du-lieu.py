#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Trích xuất danh mục thủ tục hành chính từ file Excel nguồn thành các file JSON
mà website đọc lúc build.

Cách dùng:
    python3 scripts/trich-xuat-du-lieu.py data/source/DANH_MUC_TTHC.xlsx

Nếu không truyền đường dẫn, script tự tìm file .xlsx đầu tiên trong data/source/.

Đầu ra (ghi vào data/):
    tthc.json      - danh sách TTHC, đã ghép các bảng con
    linh-vuc.json  - nhóm theo lĩnh vực, kèm slug dùng cho URL và tên file QR
    meta.json      - số liệu tổng hợp, hiển thị ở trang chủ và chân trang
"""
from __future__ import annotations

import argparse
import json
import re
import sys
import unicodedata
from collections import defaultdict
from datetime import date
from pathlib import Path
from typing import Any, Iterable

import openpyxl

GOC_DU_AN = Path(__file__).resolve().parent.parent
THU_MUC_NGUON = GOC_DU_AN / "data" / "source"
THU_MUC_RA = GOC_DU_AN / "data"


def tao_slug(text: Any) -> str:
    """Chuyển tên tiếng Việt thành slug URL - phải khớp taoSlug() trong src/lib/text.ts."""
    if not text:
        return ""
    text = str(text).strip().lower()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(c for c in text if not unicodedata.combining(c))
    text = text.replace("đ", "d").replace("Đ", "d")
    text = re.sub(r"[^a-z0-9]+", "-", text)
    return re.sub(r"-+", "-", text).strip("-")


def lam_sach(v: Any) -> Any:
    if v is None:
        return ""
    return v.strip() if isinstance(v, str) else v


def gom_theo_ma(wb, ten_sheet: str, cot_ma: int, dong_bat_dau: int = 2):
    """Gom các dòng của một sheet phụ theo mã TTHC. Bỏ qua nếu sheet không tồn tại."""
    if ten_sheet not in wb.sheetnames:
        print(f"  ! Bỏ qua sheet không có trong file: {ten_sheet}", file=sys.stderr)
        return {}
    ws = wb[ten_sheet]
    nhom: dict[str, list] = defaultdict(list)
    for r in ws.iter_rows(min_row=dong_bat_dau, values_only=True):
        if not r or len(r) <= cot_ma or not r[cot_ma]:
            continue
        nhom[lam_sach(r[cot_ma])].append(r)
    return nhom


def gan_bang_con(theo_ma, nhom, khoa: str, anh_xa) -> None:
    for ma, cac_dong in nhom.items():
        if ma in theo_ma:
            theo_ma[ma][khoa] = [anh_xa(r) for r in cac_dong]


def o(r: Iterable, i: int) -> Any:
    """Lấy ô thứ i của dòng, trả về "" nếu dòng ngắn hơn - tránh IndexError khi
    file Excel nguồn thiếu cột."""
    r = list(r)
    return lam_sach(r[i]) if len(r) > i else ""


def tim_file_nguon(tham_so: str | None) -> Path | None:
    if tham_so:
        p = Path(tham_so)
        return p if p.is_absolute() else GOC_DU_AN / p
    if THU_MUC_NGUON.is_dir():
        ung_vien = sorted(THU_MUC_NGUON.glob("*.xlsx"))
        if ung_vien:
            return ung_vien[0]
    return None


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("nguon", nargs="?", help="Đường dẫn file Excel danh mục TTHC")
    parser.add_argument(
        "--ngay-xuat",
        default=date.today().strftime("%d/%m/%Y"),
        help="Ngày xuất dữ liệu hiển thị trên website (mặc định: hôm nay)",
    )
    args = parser.parse_args()

    src = tim_file_nguon(args.nguon)
    if src is None or not src.exists():
        print(
            "LỖI: không tìm thấy file Excel nguồn.\n"
            f"Hãy đặt file .xlsx vào {THU_MUC_NGUON}/ hoặc truyền đường dẫn:\n"
            "    python3 scripts/trich-xuat-du-lieu.py duong/dan/file.xlsx",
            file=sys.stderr,
        )
        return 1

    print(f"Đọc: {src}")
    wb = openpyxl.load_workbook(src, data_only=True)

    # ---------- 1. Danh mục TTHC chính (cấp xã) ----------
    ws = wb["TTHC2-CAPXA"]
    danh_sach: list[dict] = []
    for r in ws.iter_rows(min_row=4, values_only=True):
        if not r or not r[0] or r[0] == "STT" or not lam_sach(r[1]):
            continue
        danh_sach.append(
            {
                "stt": o(r, 0),
                "ma_tthc": o(r, 1),
                "ten_tthc": o(r, 2),
                "link_truy_cap": o(r, 3),
                "so_quyet_dinh": o(r, 4),
                "ngay_quyet_dinh": o(r, 5),
                "co_quan_ban_hanh": o(r, 6),
                "linh_vuc": o(r, 7),
                "cap_thuc_hien": o(r, 8),
                "doi_tuong_thuc_hien": o(r, 9),
                "loai_tthc": o(r, 10),
                "tu_khoa": o(r, 11),
                "dia_chi_tiep_nhan": o(r, 12),
                "co_quan_thuc_hien": o(r, 13),
                "co_quan_co_tham_quyen": o(r, 14),
                "co_quan_phoi_hop": o(r, 15),
                "co_quan_duoc_uy_quyen": o(r, 16),
                "mo_ta": o(r, 17),
                "trang_thai": o(r, 18),
            }
        )

    print(f"  TTHC: {len(danh_sach)}")
    theo_ma = {t["ma_tthc"]: t for t in danh_sach}

    # ---------- 2. Các bảng con, ghép theo mã TTHC ----------
    gan_bang_con(
        theo_ma,
        gom_theo_ma(wb, "Cách thức thực hiện", 1),
        "cach_thuc_thuc_hien",
        lambda r: {
            "hinh_thuc_nop": o(r, 3),
            "thoi_gian": o(r, 4),
            "phi_le_phi": o(r, 5),
            "mo_ta": o(r, 6),
        },
    )
    gan_bang_con(
        theo_ma,
        gom_theo_ma(wb, "Trình tự thực hiện", 1),
        "trinh_tu_thuc_hien",
        lambda r: {"truong_hop_xu_ly": o(r, 3), "mo_ta_chi_tiet": o(r, 4)},
    )
    gan_bang_con(
        theo_ma,
        gom_theo_ma(wb, "TPHS", 1),
        "tphs",
        lambda r: {
            "ma_tphs": o(r, 3),
            "ten_giay_to": o(r, 4),
            "loai_giay_to": o(r, 5),
            "co_quan_ban_hanh": o(r, 6),
            "ban_chinh": bool(o(r, 7)),
            "ban_sao": bool(o(r, 8)),
            "bm_dien_tu": o(r, 9),
        },
    )
    gan_bang_con(
        theo_ma,
        gom_theo_ma(wb, "KQTH", 1),
        "kqth",
        lambda r: {
            "ma_ket_qua": o(r, 3),
            "ten_ket_qua": o(r, 4),
            "co_quan_ban_hanh": o(r, 5),
            "mo_ta": o(r, 6),
        },
    )
    gan_bang_con(
        theo_ma,
        gom_theo_ma(wb, "CCPL", 1),
        "ccpl",
        lambda r: {
            "so_ky_hieu": o(r, 3),
            "trich_yeu": o(r, 4),
            "co_quan_ban_hanh": o(r, 5),
            "ngay_ban_hanh": o(r, 6),
            "ngay_hieu_luc": o(r, 7),
            "dia_chi_truy_cap": o(r, 8),
        },
    )

    # ---------- 3. Yêu cầu, điều kiện (sheet TTHC toàn quốc, khớp theo mã) ----------
    if "TTHC" in wb.sheetnames:
        for r in wb["TTHC"].iter_rows(min_row=2, values_only=True):
            if not r or not r[1]:
                continue
            ma = lam_sach(r[1])
            if ma in theo_ma:
                theo_ma[ma]["yeu_cau_dieu_kien"] = o(r, 16)

    # ---------- 4. Nhóm theo lĩnh vực ----------
    theo_linh_vuc: dict[str, list[str]] = defaultdict(list)
    for t in danh_sach:
        theo_linh_vuc[t["linh_vuc"] or "Chưa phân loại"].append(t["ma_tthc"])

    linh_vuc = [
        {
            "ten_linh_vuc": ten,
            "slug": tao_slug(ten),
            "so_luong_tthc": len(ma_list),
            "danh_sach_ma_tthc": ma_list,
        }
        for ten, ma_list in sorted(theo_linh_vuc.items())
    ]
    print(f"  Lĩnh vực: {len(linh_vuc)}")

    # Slug trùng nhau sẽ khiến hai lĩnh vực dùng chung một URL và một mã QR.
    slugs = [lv["slug"] for lv in linh_vuc]
    trung = {s for s in slugs if slugs.count(s) > 1}
    if trung:
        print(f"LỖI: slug lĩnh vực bị trùng: {trung}", file=sys.stderr)
        return 1

    # ---------- 5. Ghi JSON ----------
    THU_MUC_RA.mkdir(parents=True, exist_ok=True)

    def ghi(ten: str, du_lieu, thut_le=None):
        duong_dan = THU_MUC_RA / ten
        duong_dan.write_text(
            json.dumps(du_lieu, ensure_ascii=False, indent=thut_le), encoding="utf-8"
        )
        print(f"  → {ten} ({duong_dan.stat().st_size / 1024:.0f} KB)")

    ghi("tthc.json", danh_sach)
    ghi("linh-vuc.json", linh_vuc)
    ghi(
        "meta.json",
        {
            "tong_so_tthc": len(danh_sach),
            "tong_so_linh_vuc": len(linh_vuc),
            "ngay_xuat": args.ngay_xuat,
            "nguon": "Cổng Dịch vụ công Quốc gia (dichvucong.gov.vn)",
        },
        thut_le=2,
    )

    print("Xong. Chạy tiếp: python3 scripts/tao-ma-qr.py")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
